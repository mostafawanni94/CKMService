"""
The finance dashboard and the accountant's export.

Both read the same records as the return, so the test is mainly that they agree
with it — a dashboard that contradicts the filing is worse than no dashboard.
"""

from datetime import date, timedelta
from decimal import Decimal

from django.test import TestCase
from rest_framework.test import APIClient

from apps.core.models import SystemConfig
from apps.core.testing import (
    attach_service_rate, make_customer, make_employee, make_project, make_service,
    make_user, make_work_entry,
)
from apps.invoices.billing import create_credit_note, generate_invoice, issue_invoice
from apps.invoices.models import IncomingInvoice, Invoice
from apps.vat.exports import build_quarter_workbook
from apps.vat.models import VatPeriod
from apps.vat.reporting import (
    costs, dashboard, monthly_series, payables, receivables, revenue, window_for,
)
from apps.vat.returns import calculate_return

MONDAY = date(2026, 8, 10)


class WindowTests(TestCase):
    def test_a_quarter(self):
        self.assertEqual(window_for(2026, quarter=3), (date(2026, 7, 1), date(2026, 9, 30)))

    def test_the_fourth_quarter_ends_the_year(self):
        self.assertEqual(window_for(2026, quarter=4), (date(2026, 10, 1), date(2026, 12, 31)))

    def test_a_whole_year(self):
        self.assertEqual(window_for(2026), (date(2026, 1, 1), date(2026, 12, 31)))

    def test_a_month(self):
        self.assertEqual(window_for(2026, month=2), (date(2026, 2, 1), date(2026, 2, 28)))


class ReportingSetup(TestCase):
    def setUp(self):
        config = SystemConfig.objects.get_config()
        config.company_legal_name = 'CKMcleaning VOF'
        config.company_btw_number = 'NL869591071B01'
        config.company_kvk_number = '42074970'
        config.save()

        self.user = make_user(email='report@ckm.test', role='admin')
        self.customer = make_customer(company_name='Smaak voor Groen',
                                      btw_number='NL001538146B17')
        self.customer.vat_treatment_code = 'NORMAL'
        self.customer.save()
        self.project = make_project(customer=self.customer)
        self.service = make_service(name='Schoonmaak')
        attach_service_rate(self.customer, self.service, Decimal('40.00'))
        self.employee = make_employee(hourly_rate=Decimal('16.00'))

    def issued_invoice(self, day=0, issue_date=None):
        make_work_entry(employee=make_employee(), project=self.project,
                        work_date=MONDAY + timedelta(days=day), service=self.service)
        invoice = generate_invoice(
            self.customer, start=MONDAY + timedelta(days=day),
            end=MONDAY + timedelta(days=day), actor=self.user)
        issue_invoice(invoice, actor=self.user, issue_date=issue_date or date(2026, 8, 17))
        invoice.refresh_from_db()
        return invoice


class RevenueTests(ReportingSetup):
    def test_only_issued_invoices_count_as_revenue(self):
        make_work_entry(employee=self.employee, project=self.project,
                        work_date=MONDAY, service=self.service)
        generate_invoice(self.customer, week=(2026, 33), actor=self.user)   # draft
        result = revenue(date(2026, 7, 1), date(2026, 9, 30))
        self.assertEqual(result['invoiced_net'], Decimal('0.00'))

    def test_an_issued_invoice_is_revenue(self):
        self.issued_invoice()
        result = revenue(date(2026, 7, 1), date(2026, 9, 30))
        self.assertEqual(result['invoiced_net'], Decimal('300.00'))
        self.assertEqual(result['invoiced_vat'], Decimal('63.00'))
        self.assertEqual(result['net_revenue'], Decimal('300.00'))

    def test_a_credit_note_reduces_revenue(self):
        invoice = self.issued_invoice()
        create_credit_note(invoice, reason='Billed to the wrong customer.',
                           actor=self.user, issue_date=date(2026, 8, 20))
        result = revenue(date(2026, 7, 1), date(2026, 9, 30))
        self.assertEqual(result['credited_net'], Decimal('300.00'))
        self.assertEqual(result['net_revenue'], Decimal('0.00'))

    def test_revenue_agrees_with_box_1a(self):
        self.issued_invoice()
        result = revenue(date(2026, 7, 1), date(2026, 9, 30))
        vat_return = calculate_return(VatPeriod.for_date(date(2026, 8, 17)))
        box_1a = next(b for b in vat_return['boxes'] if b['code'] == '1a')
        self.assertEqual(result['net_revenue'], box_1a['taxable_base'])


class CostTests(ReportingSetup):
    def test_supplier_invoices_are_costs(self):
        IncomingInvoice.objects.create(
            invoice_number='SUP-1', vendor_name='Makro', invoice_date=date(2026, 8, 5),
            subtotal=Decimal('200.00'), vat_rate=Decimal('21.00'),
            vat_amount=Decimal('42.00'), total=Decimal('242.00'))
        result = costs(date(2026, 7, 1), date(2026, 9, 30))
        self.assertEqual(result['supplier_net'], Decimal('200.00'))
        self.assertEqual(result['total_net'], Decimal('200.00'))

    def test_expenses_are_costs(self):
        from apps.expenses.models import Expense, ExpenseCategory

        category = ExpenseCategory.objects.create(name='Materiaal', code='MAT')
        Expense.objects.create(category=category, description='Doeken',
                               vendor_name='Makro', amount_excl_vat=Decimal('50.00'),
                               vat_rate=Decimal('21.00'), expense_date=date(2026, 8, 5))
        result = costs(date(2026, 7, 1), date(2026, 9, 30))
        self.assertEqual(result['expense_net'], Decimal('50.00'))


class ReceivablesTests(ReportingSetup):
    def test_an_unpaid_invoice_is_outstanding(self):
        self.issued_invoice()
        result = receivables(as_of=date(2026, 8, 20))
        self.assertEqual(result['total_outstanding'], Decimal('363.00'))
        self.assertEqual(result['overdue_count'], 0)

    def test_ageing_buckets_by_days_overdue(self):
        self.issued_invoice(issue_date=date(2026, 6, 1))   # due 15 June
        result = receivables(as_of=date(2026, 8, 20))
        self.assertEqual(result['overdue_count'], 1)
        self.assertEqual(result['ageing']['days_61_90'], Decimal('363.00'))

    def test_a_paid_invoice_drops_out(self):
        from apps.invoices.billing import record_payment

        invoice = self.issued_invoice()
        record_payment(invoice, invoice.total)
        self.assertEqual(receivables()['total_outstanding'], Decimal('0.00'))


class PayablesTests(ReportingSetup):
    def test_wallet_balances_are_a_liability(self):
        from apps.wallet.services import credit_work_entry

        entry = make_work_entry(employee=self.employee, project=self.project,
                                work_date=MONDAY, service=self.service)
        credit_work_entry(entry)
        result = payables()
        self.assertEqual(result['employee_wallets'], Decimal('120.00'))
        self.assertEqual(result['total'], Decimal('120.00'))

    def test_unpaid_supplier_invoices_are_a_liability(self):
        IncomingInvoice.objects.create(
            invoice_number='SUP-2', vendor_name='Makro', invoice_date=date(2026, 8, 5),
            subtotal=Decimal('200.00'), vat_rate=Decimal('21.00'),
            vat_amount=Decimal('42.00'), total=Decimal('242.00'))
        self.assertEqual(payables()['supplier_outstanding'], Decimal('242.00'))


class DashboardTests(ReportingSetup):
    def test_the_dashboard_reports_the_quarter(self):
        self.issued_invoice()
        result = dashboard(2026, quarter=3)
        self.assertEqual(result['revenue']['net_revenue'], Decimal('300.00'))
        self.assertEqual(result['gross_margin'], Decimal('300.00'))
        self.assertEqual(len(result['monthly']), 12)

    def test_the_dashboard_vat_agrees_with_the_return(self):
        self.issued_invoice()
        result = dashboard(2026, quarter=3)
        period = VatPeriod.for_date(date(2026, 8, 17))
        self.assertEqual(result['vat_return']['box_5a'],
                         calculate_return(period)['box_5a'])

    def test_unresolved_transactions_are_surfaced(self):
        IncomingInvoice.objects.create(
            invoice_number='654646', vendor_name='8776', invoice_date=date(2026, 8, 5),
            subtotal=Decimal('3000.00'), vat_rate=Decimal('21.00'),
            vat_treatment_code='NORMAL')
        from apps.vat.posting import post_all
        post_all()

        result = dashboard(2026, quarter=3)
        self.assertEqual(result['requires_review_count'], 1)

    def test_monthly_series_covers_every_month(self):
        self.issued_invoice()
        series = monthly_series(2026)
        august = next(row for row in series if row['month'] == 8)
        self.assertEqual(august['revenue'], Decimal('300.00'))
        self.assertEqual(sum(1 for row in series if row['revenue'] == 0), 11)


class ExportTests(ReportingSetup):
    def test_the_workbook_is_produced(self):
        self.issued_invoice()
        content = build_quarter_workbook(VatPeriod.for_date(date(2026, 8, 17)))
        self.assertTrue(content.startswith(b'PK'))       # xlsx is a zip
        self.assertGreater(len(content), 4000)

    def test_the_workbook_carries_the_return_and_the_transactions(self):
        from io import BytesIO

        from openpyxl import load_workbook

        self.issued_invoice()
        period = VatPeriod.for_date(date(2026, 8, 17))
        book = load_workbook(BytesIO(build_quarter_workbook(period)))
        self.assertIn('Aangifte', book.sheetnames)
        self.assertIn('Transacties', book.sheetnames)
        self.assertIn('Documenten', book.sheetnames)

        values = [row for row in book['Aangifte'].values]
        flat = [str(cell) for row in values for cell in row if cell is not None]
        self.assertIn('1a', flat)
        self.assertNotIn('5g', flat)

    def test_unresolved_transactions_get_their_own_sheet(self):
        from io import BytesIO

        from openpyxl import load_workbook

        IncomingInvoice.objects.create(
            invoice_number='654646', vendor_name='8776', invoice_date=date(2026, 8, 5),
            subtotal=Decimal('3000.00'), vat_rate=Decimal('21.00'),
            vat_treatment_code='NORMAL')
        from apps.vat.posting import post_all
        post_all()

        period = VatPeriod.for_date(date(2026, 8, 5))
        book = load_workbook(BytesIO(build_quarter_workbook(period)))
        self.assertIn('Vast te stellen', book.sheetnames)


class FinanceApiTests(ReportingSetup):
    def setUp(self):
        super().setUp()
        self.client = APIClient()
        self.client.force_authenticate(make_user(email='fd@ckm.test', role='finance'))

    def test_the_dashboard_endpoint(self):
        self.issued_invoice()
        response = self.client.get('/api/vat/dashboard/?year=2026&quarter=3')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data['revenue']['net_revenue'], Decimal('300.00'))

    def test_the_receivables_and_payables_endpoints(self):
        self.assertEqual(self.client.get('/api/vat/dashboard/receivables/').status_code, 200)
        self.assertEqual(self.client.get('/api/vat/dashboard/payables/').status_code, 200)

    def test_the_review_work_list(self):
        response = self.client.get('/api/vat/dashboard/requires-review/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('entries', response.data)

    def test_the_export_endpoint_returns_a_workbook(self):
        self.issued_invoice()
        period = VatPeriod.for_date(date(2026, 8, 17))
        response = self.client.get(f'/api/vat/periods/{period.pk}/export/')
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])
        self.assertIn('Aangifte_2026_Q3.xlsx', response['Content-Disposition'])

    def test_other_roles_are_refused(self):
        for role in ('operations', 'employee', 'customer'):
            client = APIClient()
            client.force_authenticate(make_user(email=f'{role}@fd.test', role=role))
            self.assertEqual(client.get('/api/vat/dashboard/').status_code, 403)
