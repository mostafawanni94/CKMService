"""
Exports for the accountant.

A bookkeeper needs three things from a quarter: the return itself, the
transactions behind every box, and the source documents. This module produces
the first two as one Excel workbook whose figures come from the same calculator
the filing uses, so what the accountant checks is what would be filed.

Nothing here recomputes VAT.
"""

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .returns import calculate_return, entries_for_box

HEADER_FILL = PatternFill('solid', fgColor='1E3A5F')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color='1E3A5F')
LABEL_FONT = Font(bold=True, size=10)
MONEY = '#,##0.00'
THIN = Side(style='thin', color='D9E2EC')
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_header(sheet, row, headers, widths=None):
    for column, title in enumerate(headers, start=1):
        cell = sheet.cell(row=row, column=column, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = BOX
    for column, width in enumerate(widths or [], start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = sheet.cell(row=row + 1, column=1)
    return row + 1


def _money(sheet, row, column, value):
    cell = sheet.cell(row=row, column=column, value=float(value or 0))
    cell.number_format = MONEY
    return cell


def _return_sheet(workbook, period, result, config):
    sheet = workbook.active
    sheet.title = 'Aangifte'

    sheet['A1'] = f'Btw-aangifte {period}'
    sheet['A1'].font = TITLE_FONT
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 52
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 10

    identity = [
        ('Onderneming', config.company_legal_name or config.company_name),
        ('Btw-nummer', config.company_btw_number),
        ('KvK-nummer', config.company_kvk_number),
        ('Tijdvak', f"{result['start_date']} t/m {result['end_date']}"),
        ('Status', result['status']),
        ('Regelversie', result['rules_version']),
        ('Berekend op', str(result['calculated_at'])[:19]),
        ('Stelsel', 'Factuurstelsel — het tijdvak volgt de factuurdatum, '
                    'niet de betaaldatum'),
    ]
    row = 3
    for label, value in identity:
        sheet.cell(row=row, column=1, value=label).font = LABEL_FONT
        sheet.cell(row=row, column=2, value=str(value or ''))
        row += 1

    row += 1
    row = _write_header(sheet, row, ['Rubriek', 'Omschrijving', 'Bedrag',
                                     'Omzetbelasting', 'Posten'])
    for box in result['boxes']:
        sheet.cell(row=row, column=1, value=box['code'])
        sheet.cell(row=row, column=2, value=box['name'])
        _money(sheet, row, 3, box['taxable_base'])
        _money(sheet, row, 4, box['vat_amount'])
        sheet.cell(row=row, column=5, value=box['entry_count'])
        row += 1

    row += 1
    outcome = {'PAYABLE': 'Te betalen', 'REFUNDABLE': 'Terug te ontvangen',
               'ZERO': 'Saldo nihil'}[result['outcome']]
    for label, value in (('5a — Verschuldigde omzetbelasting', result['box_5a']),
                         ('5b — Voorbelasting', result['box_5b']),
                         (outcome, abs(result['vat_position']))):
        sheet.cell(row=row, column=2, value=label).font = LABEL_FONT
        cell = _money(sheet, row, 4, value)
        cell.font = LABEL_FONT
        row += 1

    if result['requires_review_count']:
        row += 1
        warning = sheet.cell(
            row=row, column=2,
            value=f"Let op: {result['requires_review_count']} transactie(s) zijn "
                  f"nog niet vastgesteld en zitten NIET in deze aangifte.")
        warning.font = Font(bold=True, color='B45309')
    return sheet


def _entries_sheet(workbook, period, result):
    sheet = workbook.create_sheet('Transacties')
    row = _write_header(
        sheet, 1,
        ['Rubriek', 'Datum', 'Bron', 'Referentie', 'Behandeling', 'Tarief',
         'Grondslag', 'Btw', 'Aftrekbaar', 'Status'],
        [10, 12, 18, 26, 20, 8, 14, 14, 14, 18])

    for box in result['boxes']:
        if not box['entry_count']:
            continue
        for entry in entries_for_box(period, box['code']).select_related('treatment'):
            sheet.cell(row=row, column=1, value=box['code'])
            sheet.cell(row=row, column=2, value=entry.transaction_date)
            sheet.cell(row=row, column=3, value=entry.source_type)
            sheet.cell(row=row, column=4, value=entry.source_reference or '')
            sheet.cell(row=row, column=5, value=entry.treatment_code)
            sheet.cell(row=row, column=6, value=float(entry.vat_rate or 0))
            _money(sheet, row, 7, entry.taxable_base)
            _money(sheet, row, 8, entry.vat_amount)
            _money(sheet, row, 9, entry.deductible_vat)
            sheet.cell(row=row, column=10, value=entry.classification_status)
            row += 1
    return sheet


def _review_sheet(workbook, period):
    """Everything the engine refused to decide, and why."""
    from .constants import ClassificationStatus
    from .models import VatLedgerEntry

    unresolved = VatLedgerEntry.objects.filter(
        period=period, is_deleted=False,
        classification_status=ClassificationStatus.REQUIRES_REVIEW,
    ).order_by('transaction_date')
    if not unresolved.exists():
        return None

    sheet = workbook.create_sheet('Vast te stellen')
    row = _write_header(
        sheet, 1,
        ['Datum', 'Bron', 'Referentie', 'Documentbedrag', 'In aangifte', 'Reden'],
        [12, 18, 26, 16, 14, 90])
    for entry in unresolved:
        sheet.cell(row=row, column=1, value=entry.transaction_date)
        sheet.cell(row=row, column=2, value=entry.source_type)
        sheet.cell(row=row, column=3, value=entry.source_reference or '')
        _money(sheet, row, 4, entry.taxable_base)
        _money(sheet, row, 5, entry.vat_amount)
        cell = sheet.cell(row=row, column=6, value=entry.review_reason or '')
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    return sheet


def _documents_sheet(workbook, period):
    """The invoices and costs of the quarter, as the bookkeeper's own list."""
    from apps.expenses.models import Expense
    from apps.invoices.models import AgencyInvoice, IncomingInvoice, Invoice

    sheet = workbook.create_sheet('Documenten')
    row = _write_header(
        sheet, 1,
        ['Soort', 'Nummer', 'Datum', 'Tegenpartij', 'Netto', 'Btw', 'Totaal',
         'Status'],
        [16, 18, 12, 34, 14, 14, 14, 16])

    start, end = period.start_date, period.end_date

    for invoice in Invoice.objects.filter(
            is_deleted=False, issue_date__gte=start, issue_date__lte=end
    ).select_related('customer').order_by('issue_date'):
        sheet.cell(row=row, column=1,
                   value='Creditnota' if invoice.is_credit_note else 'Verkoopfactuur')
        sheet.cell(row=row, column=2, value=invoice.invoice_number)
        sheet.cell(row=row, column=3, value=invoice.issue_date)
        sheet.cell(row=row, column=4, value=invoice.customer.company_name)
        _money(sheet, row, 5, invoice.subtotal)
        _money(sheet, row, 6, invoice.vat_amount)
        _money(sheet, row, 7, invoice.total)
        sheet.cell(row=row, column=8, value=invoice.get_status_display())
        row += 1

    for incoming in IncomingInvoice.objects.filter(
            is_deleted=False, invoice_date__gte=start, invoice_date__lte=end
    ).order_by('invoice_date'):
        sheet.cell(row=row, column=1, value='Inkoopfactuur')
        sheet.cell(row=row, column=2, value=incoming.invoice_number)
        sheet.cell(row=row, column=3, value=incoming.invoice_date)
        sheet.cell(row=row, column=4, value=incoming.vendor_name)
        _money(sheet, row, 5, incoming.subtotal)
        _money(sheet, row, 6, incoming.vat_amount)
        _money(sheet, row, 7, incoming.total)
        sheet.cell(row=row, column=8, value=incoming.get_status_display())
        row += 1

    for agency in AgencyInvoice.objects.filter(
            is_deleted=False, issue_date__gte=start, issue_date__lte=end
    ).select_related('agency').order_by('issue_date'):
        sheet.cell(row=row, column=1, value='Uitzendfactuur')
        sheet.cell(row=row, column=2, value=agency.invoice_number)
        sheet.cell(row=row, column=3, value=agency.issue_date)
        sheet.cell(row=row, column=4, value=agency.agency.name)
        _money(sheet, row, 5, agency.subtotal + agency.total_surcharges)
        _money(sheet, row, 6, agency.vat_amount)
        _money(sheet, row, 7, agency.total)
        sheet.cell(row=row, column=8, value=agency.get_status_display())
        row += 1

    for expense in Expense.objects.filter(
            is_deleted=False, expense_date__gte=start, expense_date__lte=end
    ).select_related('category').order_by('expense_date'):
        sheet.cell(row=row, column=1, value='Kosten')
        sheet.cell(row=row, column=2, value=expense.reference_number or '')
        sheet.cell(row=row, column=3, value=expense.expense_date)
        sheet.cell(row=row, column=4, value=expense.vendor_name)
        _money(sheet, row, 5, expense.amount_excl_vat)
        _money(sheet, row, 6, expense.vat_amount)
        _money(sheet, row, 7, expense.total_amount)
        sheet.cell(row=row, column=8, value=expense.get_status_display())
        row += 1

    return sheet


def build_quarter_workbook(period):
    """
    The accountant's file for one quarter.

    Four sheets: the return, the transactions behind every box, anything the
    engine refused to decide, and the documents of the quarter.
    """
    from apps.core.models import SystemConfig

    config = SystemConfig.objects.get_config()
    result = calculate_return(period)

    workbook = Workbook()
    _return_sheet(workbook, period, result, config)
    _entries_sheet(workbook, period, result)
    _review_sheet(workbook, period)
    _documents_sheet(workbook, period)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def workbook_filename(period):
    return f'Aangifte_{period.year}_Q{period.quarter}.xlsx'
