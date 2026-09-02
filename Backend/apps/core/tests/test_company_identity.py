"""
The company's legal identity.

A Dutch invoice must carry the supplier's name, address, KvK number and
BTW-identificatienummer. These live on one record — `SystemConfig` — and every
document reads them from there, so a correction is made once and appears
everywhere.
"""

from datetime import date
from decimal import Decimal
from io import StringIO

from django.core.management import call_command
from django.test import TestCase
from rest_framework.test import APIClient

from apps.core.models import SystemConfig
from apps.core.testing import (
    attach_service_rate, make_customer, make_employee, make_project, make_service,
    make_user, make_work_entry,
)

REGISTERED = {
    'company_legal_name': 'CKM Services',
    'company_address': 'Rilland Bathstraat 126',
    'company_postal_code': '3086 ST',
    'company_city': 'Rotterdam',
    'company_country': 'Netherlands',
    'company_kvk_number': '42074970',
    'company_btw_number': 'NL869591071B01',
    'company_iban': 'NL20 INGB 0119 4132 56',
}


class ConfigurationTests(TestCase):
    def test_the_command_writes_the_registered_values(self):
        call_command('set_company_identity', stdout=StringIO())

        config = SystemConfig.objects.get_config()
        for field, expected in REGISTERED.items():
            with self.subTest(field=field):
                self.assertEqual(getattr(config, field), expected)

    def test_the_primary_email_is_recorded(self):
        call_command('set_company_identity', stdout=StringIO())
        config = SystemConfig.objects.get_config()
        addresses = [row.get('email') for row in config.company_emails]
        self.assertIn('info@ckmservices.nl', addresses)

    def test_running_it_twice_changes_nothing(self):
        call_command('set_company_identity', stdout=StringIO())
        out = StringIO()
        call_command('set_company_identity', stdout=out)
        self.assertIn('Already up to date', out.getvalue())

    def test_there_is_only_ever_one_configuration_record(self):
        call_command('set_company_identity', stdout=StringIO())
        call_command('set_company_identity', stdout=StringIO())
        self.assertEqual(SystemConfig.objects.count(), 1)

    def test_show_masks_the_iban(self):
        call_command('set_company_identity', stdout=StringIO())
        out = StringIO()
        call_command('set_company_identity', '--show', stdout=out)
        printed = out.getvalue()
        self.assertIn('company_kvk_number', printed)
        self.assertIn('42074970', printed)
        self.assertNotIn('INGB 0119 4132', printed, 'the IBAN was printed in full')


class StorageTests(TestCase):
    def setUp(self):
        call_command('set_company_identity', stdout=StringIO())

    def test_the_iban_is_ciphertext_at_rest(self):
        from django.db import connection

        config = SystemConfig.objects.get_config()
        with connection.cursor() as cursor:
            cursor.execute('SELECT company_iban FROM core_systemconfig WHERE id = %s',
                           [config.pk])
            stored = cursor.fetchone()[0]

        self.assertTrue(stored.startswith('enc$v1'))
        self.assertNotIn('INGB', stored)
        self.assertEqual(config.company_iban, 'NL20 INGB 0119 4132 56')

    def test_the_public_endpoint_does_not_carry_it(self):
        response = APIClient().get('/api/settings/config/public/')
        self.assertEqual(response.status_code, 200)
        self.assertNotIn('company_iban', response.data)
        self.assertNotIn('INGB', str(response.data))

    def test_a_non_admin_does_not_see_it(self):
        client = APIClient()
        client.force_authenticate(make_employee().user)
        response = client.get('/api/settings/config/')
        self.assertNotIn('company_iban', response.data)

    def test_an_admin_may_read_and_change_it(self):
        client = APIClient()
        client.force_authenticate(make_user(email='ci-admin@ckm.test', role='admin'))
        response = client.get('/api/settings/config/')
        self.assertEqual(response.data['company_kvk_number'], '42074970')
        self.assertEqual(response.data['company_iban'], 'NL20 INGB 0119 4132 56')


class ContactShapeTests(TestCase):
    """
    Two shapes for a contact entry are in the wild.

    The settings page writes {'label', 'value'}; the model's help text and the
    invoice template expect {'label', 'email'} / {'label', 'number'}. The PDF
    read only the latter, so an address or phone typed into Settings never
    appeared on an invoice. The accessors take either.
    """

    def setUp(self):
        self.config = SystemConfig.objects.get_config()

    def test_the_settings_page_shape_is_read(self):
        self.config.company_emails = [{'label': 'Main', 'value': 'a@ckm.test'}]
        self.config.company_phones = [{'label': 'Main', 'value': '+31 6 12345678'}]
        self.config.save()

        self.assertEqual(self.config.contact_emails, ['a@ckm.test'])
        self.assertEqual(self.config.contact_phones, ['+31 6 12345678'])

    def test_the_documented_shape_is_read(self):
        self.config.company_emails = [{'label': 'Info', 'email': 'b@ckm.test'}]
        self.config.company_phones = [{'label': 'Main', 'number': '+31 6 87654321'}]
        self.config.save()

        self.assertEqual(self.config.contact_emails, ['b@ckm.test'])
        self.assertEqual(self.config.contact_phones, ['+31 6 87654321'])

    def test_both_shapes_together_do_not_duplicate_an_address(self):
        self.config.company_emails = [
            {'label': 'Info', 'email': 'same@ckm.test'},
            {'label': 'Main', 'value': 'same@ckm.test'},
        ]
        self.config.save()
        self.assertEqual(self.config.contact_emails, ['same@ckm.test'])

    def test_junk_entries_are_ignored_rather_than_crashing_an_invoice(self):
        self.config.company_emails = [{}, {'label': 'Empty', 'email': ''}, None,
                                      {'label': 'Good', 'value': 'c@ckm.test'}]
        self.config.save()
        self.assertEqual(self.config.contact_emails, ['c@ckm.test'])

    def test_the_command_does_not_add_a_second_copy(self):
        self.config.company_emails = [{'label': 'Main', 'value': 'info@ckmservices.nl'}]
        self.config.save()

        call_command('set_company_identity', stdout=StringIO())
        self.config.refresh_from_db()
        self.assertEqual(self.config.contact_emails, ['info@ckmservices.nl'])


class DocumentTests(TestCase):
    """The identity reaches every document that legally needs it."""

    def setUp(self):
        call_command('set_company_identity', stdout=StringIO())
        self.user = make_user(email='ci-doc@ckm.test', role='admin')
        self.customer = make_customer(company_name='Smaak voor Groen',
                                      btw_number='NL001538146B17')
        self.customer.vat_treatment_code = 'NORMAL'
        self.customer.save()
        self.project = make_project(customer=self.customer)
        self.service = make_service(name='Schoonmaak')
        attach_service_rate(self.customer, self.service, Decimal('40.00'))
        make_work_entry(employee=make_employee(), project=self.project,
                        service=self.service, work_date=date(2026, 8, 10))

        from apps.invoices.billing import generate_invoice, issue_invoice

        self.invoice = generate_invoice(self.customer, week=(2026, 33),
                                        actor=self.user)
        issue_invoice(self.invoice, actor=self.user, issue_date=date(2026, 8, 17))
        self.invoice.refresh_from_db()

    def _text(self, document):
        import io

        from PyPDF2 import PdfReader

        from apps.invoices.pdf import build_invoice_pdf

        return PdfReader(io.BytesIO(build_invoice_pdf(document))).pages[0].extract_text()

    def test_the_invoice_carries_everything_the_law_requires(self):
        text = self._text(self.invoice)
        for required in ('CKM Services', 'Rilland Bathstraat 126', '3086 ST',
                         'Rotterdam', 'KvK 42074970', 'BTW NL869591071B01',
                         'NL20 INGB 0119 4132 56'):
            with self.subTest(required=required):
                self.assertIn(required, text, f'the invoice omits {required}')

    def test_the_contact_details_reach_the_invoice(self):
        config = SystemConfig.objects.get_config()
        config.company_phones = [{'label': 'Main', 'value': '+31 6 84466318'}]
        config.save()

        text = self._text(self.invoice)
        self.assertIn('info@ckmservices.nl', text)
        self.assertIn('+31 6 84466318', text,
                      'a phone entered through Settings is missing from the invoice')

    def test_the_payment_instruction_names_the_account_and_the_company(self):
        text = self._text(self.invoice)
        self.assertIn('over te maken op NL20 INGB 0119 4132 56', text)
        self.assertIn('name van CKM Services', text)

    def test_a_credit_note_carries_the_same_identity(self):
        from apps.invoices.billing import create_credit_note

        note = create_credit_note(self.invoice, reason='A written reason for it.',
                                  actor=self.user, issue_date=date(2026, 8, 20))
        text = self._text(note)
        self.assertIn('CREDITNOTA', text)
        for required in ('CKM Services', 'KvK 42074970', 'BTW NL869591071B01'):
            with self.subTest(required=required):
                self.assertIn(required, text)

    def test_the_accountant_export_names_the_company(self):
        from openpyxl import load_workbook
        from io import BytesIO

        from apps.vat.exports import build_quarter_workbook
        from apps.vat.models import VatPeriod

        period = VatPeriod.for_date(date(2026, 8, 17))
        book = load_workbook(BytesIO(build_quarter_workbook(period)))
        values = [str(cell) for row in book['Aangifte'].values
                  for cell in row if cell is not None]

        self.assertIn('CKM Services', values)
        self.assertIn('NL869591071B01', values)
        self.assertIn('42074970', values)

    def test_the_accountant_export_does_not_carry_the_iban(self):
        """It goes to a third party; the account number is not needed there."""
        from openpyxl import load_workbook
        from io import BytesIO

        from apps.vat.exports import build_quarter_workbook
        from apps.vat.models import VatPeriod

        period = VatPeriod.for_date(date(2026, 8, 17))
        book = load_workbook(BytesIO(build_quarter_workbook(period)))
        for sheet in book.worksheets:
            for row in sheet.values:
                for cell in row:
                    if cell is not None:
                        self.assertNotIn('INGB', str(cell))

    def test_issuing_does_not_depend_on_a_hardcoded_identity(self):
        """
        Clearing the configuration must change the document, proving the PDF
        reads the record rather than a value baked into the code.
        """
        config = SystemConfig.objects.get_config()
        config.company_kvk_number = '99999999'
        config.save()

        self.assertIn('KvK 99999999', self._text(self.invoice))
