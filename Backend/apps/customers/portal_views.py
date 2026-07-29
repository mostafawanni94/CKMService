"""
Customer Portal Views - Read-only API for customer mobile app.

All endpoints are scoped to the authenticated customer user's company.
No financial data, no personal employee data (BSN, address, salary) is exposed.
"""

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count

from apps.projects.models import Project
from apps.worklogs.models import WorkEntry
from .models import Customer
from .portal_serializers import (
    CustomerProfileSerializer,
    CustomerProjectListSerializer,
    CustomerProjectDetailSerializer,
    CustomerWorkEntrySerializer,
)


# =============================================================================
# PERMISSION
# =============================================================================

class IsCustomerUser(permissions.BasePermission):
    """
    Only allow authenticated users with role='customer' and a linked customer.
    """
    message = 'Only customer portal users can access this endpoint.'
    
    def has_permission(self, request, view):
        return (
            request.user
            and request.user.is_authenticated
            and request.user.role == 'customer'
            and request.user.customer is not None
        )


# =============================================================================
# CUSTOMER PORTAL VIEWSET
# =============================================================================

class CustomerPortalViewSet(viewsets.ViewSet):
    """
    Customer Portal API — Read-only access to project data.
    
    All data is automatically scoped to the authenticated customer's company.
    
    Endpoints:
        GET /api/customer-portal/profile/              — Company profile
        GET /api/customer-portal/projects/              — List projects
        GET /api/customer-portal/projects/{id}/         — Project detail
        GET /api/customer-portal/projects/{id}/entries/ — Work entries for project
        GET /api/customer-portal/entries/{id}/          — Single work entry detail
    """
    permission_classes = [IsCustomerUser]
    
    def _get_customer(self, request):
        """Get the customer company for the authenticated user."""
        return request.user.customer
    
    # =========================================================================
    # PROFILE
    # =========================================================================
    
    @action(detail=False, methods=['get'], url_path='profile')
    def profile(self, request):
        """Get the customer's company profile."""
        customer = self._get_customer(request)
        serializer = CustomerProfileSerializer(customer, context={'request': request})
        return Response(serializer.data)
    
    # =========================================================================
    # PROJECTS
    # =========================================================================
    
    @action(detail=False, methods=['get'], url_path='projects')
    def list_projects(self, request):
        """List all projects for this customer."""
        customer = self._get_customer(request)
        
        queryset = Project.objects.filter(
            customer=customer
        ).order_by('-created_at')
        
        # Status filter
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Search
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(location__icontains=search) |
                Q(location_city__icontains=search)
            )
        
        serializer = CustomerProjectListSerializer(
            queryset, many=True, context={'request': request}
        )
        return Response({
            'count': queryset.count(),
            'results': serializer.data,
        })
    
    @action(detail=False, methods=['get'], url_path='projects/(?P<project_id>[^/.]+)')
    def project_detail(self, request, project_id=None):
        """Get detailed project info with progress and employees."""
        customer = self._get_customer(request)
        
        try:
            project = Project.objects.get(id=project_id, customer=customer)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = CustomerProjectDetailSerializer(
            project, context={'request': request}
        )
        return Response(serializer.data)
    
    # =========================================================================
    # WORK ENTRIES
    # =========================================================================
    
    @action(detail=False, methods=['get'], url_path='projects/(?P<project_id>[^/.]+)/entries')
    def project_entries(self, request, project_id=None):
        """
        List work entries for a specific project.
        
        Filters:
            - date_from: YYYY-MM-DD
            - date_to: YYYY-MM-DD
            - employee_name: partial match on first name
            - status: approved, submitted, etc.
        """
        customer = self._get_customer(request)
        
        # Verify project belongs to this customer
        try:
            project = Project.objects.get(id=project_id, customer=customer)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        queryset = WorkEntry.objects.filter(
            project=project
        ).select_related(
            'employee', 'shift_template'
        ).prefetch_related(
            'photos'
        ).exclude(
            status__in=['cancelled', 'no_show']
        ).order_by('-work_date', '-actual_start_datetime')
        
        # Apply filters
        params = request.query_params
        
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        if date_from:
            queryset = queryset.filter(work_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(work_date__lte=date_to)
        
        employee_name = params.get('employee_name')
        if employee_name:
            queryset = queryset.filter(
                Q(employee__first_name__icontains=employee_name) |
                Q(employee__last_name__icontains=employee_name)
            )
        
        status_filter = params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Work date exact filter
        work_date = params.get('work_date')
        if work_date:
            queryset = queryset.filter(work_date=work_date)
        
        # Pagination
        page = int(params.get('page', 1))
        page_size = int(params.get('page_size', 20))
        total = queryset.count()
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        
        entries = queryset[start_idx:end_idx]
        serializer = CustomerWorkEntrySerializer(
            entries, many=True, context={'request': request}
        )
        
        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'has_more': end_idx < total,
            'results': serializer.data,
        })
    
    @action(detail=False, methods=['get'], url_path='entries/(?P<entry_id>[^/.]+)')
    def entry_detail(self, request, entry_id=None):
        """Get a single work entry detail with photos."""
        customer = self._get_customer(request)
        
        try:
            entry = WorkEntry.objects.select_related(
                'employee', 'project', 'shift_template'
            ).prefetch_related(
                'photos'
            ).get(
                id=entry_id,
                project__customer=customer
            )
        except WorkEntry.DoesNotExist:
            return Response(
                {'error': 'Work entry not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = CustomerWorkEntrySerializer(
            entry, context={'request': request}
        )
        return Response(serializer.data)
    
    # =========================================================================
    # STATS / CALENDAR
    # =========================================================================
    
    @action(detail=False, methods=['get'], url_path='projects/(?P<project_id>[^/.]+)/calendar')
    def project_calendar(self, request, project_id=None):
        """
        Calendar data for a project — dates with worker counts.
        
        Response: {"days": {"2026-07-01": 3, "2026-07-02": 1, ...}}
        """
        customer = self._get_customer(request)
        
        try:
            project = Project.objects.get(id=project_id, customer=customer)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        
        qs = WorkEntry.objects.filter(
            project=project
        ).exclude(status__in=['cancelled', 'no_show'])
        
        if year:
            qs = qs.filter(work_date__year=int(year))
        if month:
            qs = qs.filter(work_date__month=int(month))
        
        counts = qs.values('work_date').annotate(
            workers_count=Count('employee', distinct=True),
            photos_count=Count('photos'),
        ).order_by('work_date')
        
        days = {}
        for row in counts:
            days[str(row['work_date'])] = {
                'workers': row['workers_count'],
                'photos': row['photos_count'],
            }
        
        return Response({
            'project_id': str(project.id),
            'project_name': project.name,
            'days': days,
        })
    
    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================
    
    @action(detail=False, methods=['get'], url_path='projects/(?P<project_id>[^/.]+)/export')
    def export_excel(self, request, project_id=None):
        """
        Export work entries to Excel for a specific project and date range.
        
        Query params:
            - date_from: YYYY-MM-DD (required)
            - date_to: YYYY-MM-DD (required)
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from zoneinfo import ZoneInfo
        from datetime import datetime
        
        customer = self._get_customer(request)
        
        try:
            project = Project.objects.get(id=project_id, customer=customer)
        except Project.DoesNotExist:
            return Response(
                {'error': 'Project not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        
        if not date_from or not date_to:
            return Response(
                {'error': 'date_from and date_to are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        amsterdam_tz = ZoneInfo('Europe/Amsterdam')
        
        entries = WorkEntry.objects.filter(
            project=project,
            work_date__gte=date_from,
            work_date__lte=date_to,
        ).select_related(
            'employee', 'shift_template'
        ).exclude(
            status__in=['cancelled', 'no_show']
        ).order_by('work_date', 'actual_start_datetime')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Work Report'
        
        # Styles
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='1B3A5C', end_color='1B3A5C', fill_type='solid')
        subheader_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
        subheader_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        data_font = Font(name='Calibri', size=11)
        bold_font = Font(name='Calibri', bold=True, size=11)
        border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB'),
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Title section
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'Work Report — {project.name}'
        title_cell.font = Font(name='Calibri', bold=True, size=16, color='1B3A5C')
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 30
        
        ws.merge_cells('A2:G2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f'Period: {date_from} to {date_to} | Customer: {customer.company_name}'
        subtitle_cell.font = Font(name='Calibri', size=11, color='6B7280')
        
        ws.merge_cells('A3:G3')
        gen_cell = ws['A3']
        gen_cell.value = f'Generated: {datetime.now(amsterdam_tz).strftime("%d/%m/%Y %H:%M")}'
        gen_cell.font = Font(name='Calibri', size=10, color='9CA3AF', italic=True)
        
        # Column headers (row 5)
        headers = ['Date', 'Employee', 'Start Time', 'End Time', 'Break (min)', 'Total Hours', 'Status']
        col_widths = [16, 22, 14, 14, 14, 14, 14]
        
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        ws.row_dimensions[5].height = 28
        
        # Data rows
        row_num = 6
        total_hours = 0
        current_date = None
        date_fill = PatternFill(start_color='F0F4FF', end_color='F0F4FF', fill_type='solid')
        alt_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        for entry in entries:
            # Date grouping — add a subtle separator when date changes
            fill = alt_fill if row_num % 2 == 0 else PatternFill(fill_type=None)
            if entry.work_date != current_date:
                current_date = entry.work_date
                fill = date_fill
            
            # Extract times
            start_str = ''
            if entry.actual_start_datetime:
                dt = entry.actual_start_datetime
                if dt.tzinfo:
                    dt = dt.astimezone(amsterdam_tz)
                start_str = dt.strftime('%H:%M')
            elif entry.planned_start_time:
                start_str = entry.planned_start_time.strftime('%H:%M')
            
            end_str = ''
            if entry.actual_end_datetime:
                dt = entry.actual_end_datetime
                if dt.tzinfo:
                    dt = dt.astimezone(amsterdam_tz)
                end_str = dt.strftime('%H:%M')
            elif entry.planned_end_time:
                end_str = entry.planned_end_time.strftime('%H:%M')
            
            hours = float(entry.calculated_hours or 0)
            total_hours += hours
            
            employee_name = ''
            if entry.employee:
                first = entry.employee.first_name or ''
                last_init = (entry.employee.last_name[0] + '.') if entry.employee.last_name else ''
                employee_name = f'{first} {last_init}'.strip()
            
            row_data = [
                entry.work_date.strftime('%d/%m/%Y'),
                employee_name,
                start_str,
                end_str,
                entry.break_duration_minutes or 0,
                hours,
                entry.get_status_display(),
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = data_font
                cell.border = border
                cell.fill = fill
                cell.alignment = center_align if col_idx >= 3 else left_align
            
            row_num += 1
        
        # Summary row
        row_num += 1
        ws.merge_cells(f'A{row_num}:E{row_num}')
        summary_cell = ws.cell(row=row_num, column=1, value=f'Total ({entries.count()} entries)')
        summary_cell.font = bold_font
        summary_cell.alignment = Alignment(horizontal='right', vertical='center')
        summary_cell.border = border
        
        total_cell = ws.cell(row=row_num, column=6, value=total_hours)
        total_cell.font = Font(name='Calibri', bold=True, size=12, color='1B3A5C')
        total_cell.border = border
        total_cell.alignment = center_align
        total_cell.number_format = '0.0'
        
        hours_label = ws.cell(row=row_num, column=7, value='hours')
        hours_label.font = bold_font
        hours_label.border = border
        hours_label.alignment = center_align
        
        # Generate response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        safe_name = project.name.replace(' ', '_')[:30]
        response['Content-Disposition'] = f'attachment; filename="WorkReport_{safe_name}_{date_from}_{date_to}.xlsx"'
        
        wb.save(response)
        return response
