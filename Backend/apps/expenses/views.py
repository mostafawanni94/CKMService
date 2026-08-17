"""
Expenses & Finance API Views.

Provides CRUD for expenses and income, plus financial summaries
and Aangifte-ready Excel export with BTW calculations.
"""
import io
from datetime import date, timedelta
from decimal import Decimal
from django.db.models import Sum, Q, F, Count
from django.db.models.functions import TruncMonth, ExtractYear
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.employees.views import IsAdmin
from .models import ExpenseCategory, Expense, IncomeRecord
from .serializers import (
    ExpenseCategorySerializer, ExpenseListSerializer,
    ExpenseDetailSerializer, ExpenseCreateSerializer,
    IncomeRecordListSerializer, IncomeRecordDetailSerializer,
    IncomeRecordCreateSerializer,
)


# =============================================================================
# EXPENSE CATEGORY VIEWSET
# =============================================================================

class ExpenseCategoryViewSet(viewsets.ModelViewSet):
    """CRUD for expense categories."""
    queryset = ExpenseCategory.objects.order_by('sort_order', 'name')
    serializer_class = ExpenseCategorySerializer
    permission_classes = [IsAdmin]


# =============================================================================
# EXPENSE VIEWSET
# =============================================================================

class ExpenseViewSet(viewsets.ModelViewSet):
    """
    CRUD for expenses with receipt upload and financial summaries.
    
    Endpoints:
        GET    /api/expenses/expenses/              — List
        POST   /api/expenses/expenses/              — Create (with receipt upload)
        GET    /api/expenses/expenses/{id}/          — Detail
        PUT    /api/expenses/expenses/{id}/          — Update
        DELETE /api/expenses/expenses/{id}/          — Delete
        GET    /api/expenses/expenses/summary/       — Financial summary
        GET    /api/expenses/expenses/export/        — Excel export for Aangifte
        POST   /api/expenses/expenses/ocr_extract/   — OCR extraction from receipt
    """
    
    queryset = Expense.objects.select_related('category').order_by('-expense_date')
    permission_classes = [IsAdmin]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return ExpenseListSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return ExpenseCreateSerializer
        return ExpenseDetailSerializer
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Category filter
        category = self.request.query_params.get('category')
        if category:
            qs = qs.filter(category_id=category)
        
        # Status filter
        status_filter = self.request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)
        
        # Date range filter
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from:
            qs = qs.filter(expense_date__gte=date_from)
        if date_to:
            qs = qs.filter(expense_date__lte=date_to)
        
        # Year filter
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(expense_date__year=int(year))
        
        # Payment method filter
        payment_method = self.request.query_params.get('payment_method')
        if payment_method:
            qs = qs.filter(payment_method=payment_method)
        
        # Search
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(description__icontains=search) |
                Q(vendor_name__icontains=search) |
                Q(reference_number__icontains=search)
            )
        
        return qs
    
    @action(detail=False, methods=['get'])
    def summary(self, request):
        """
        Financial summary with BTW overview.
        
        Query params:
            - year: year to summarize (default: current year)
            - quarter: Q1/Q2/Q3/Q4 (optional, for quarterly BTW)
        """
        year = int(request.query_params.get('year', date.today().year))
        quarter = request.query_params.get('quarter')
        
        # Build date range
        if quarter:
            q = int(quarter)
            period_start = date(year, (q - 1) * 3 + 1, 1)
            if q == 4:
                period_end = date(year, 12, 31)
            else:
                period_end = date(year, q * 3 + 1, 1) - timedelta(days=1)
        else:
            period_start = date(year, 1, 1)
            period_end = date(year, 12, 31)
        
        # Expenses summary
        expense_totals = Expense.objects.filter(
            expense_date__gte=period_start,
            expense_date__lte=period_end,
            status='approved',
        ).aggregate(
            total_excl_vat=Sum('amount_excl_vat'),
            total_vat=Sum('vat_amount'),
            total_incl_vat=Sum('total_amount'),
        )
        
        # Income summary
        income_totals = IncomeRecord.objects.filter(
            received_date__gte=period_start,
            received_date__lte=period_end,
        ).aggregate(
            total_excl_vat=Sum('amount_excl_vat'),
            total_vat=Sum('vat_amount'),
            total_incl_vat=Sum('total_amount'),
        )
        
        total_expenses = expense_totals['total_incl_vat'] or Decimal('0.00')
        total_income = income_totals['total_incl_vat'] or Decimal('0.00')
        total_vat_paid = expense_totals['total_vat'] or Decimal('0.00')
        total_vat_collected = income_totals['total_vat'] or Decimal('0.00')
        
        # Expenses by category
        by_category = Expense.objects.filter(
            expense_date__gte=period_start,
            expense_date__lte=period_end,
            status='approved',
        ).values(
            'category__name', 'category__code', 'category__color'
        ).annotate(
            total=Sum('total_amount'),
            count=Count('id'),
        ).order_by('-total')
        
        # Monthly breakdown
        monthly_expenses = Expense.objects.filter(
            expense_date__gte=period_start,
            expense_date__lte=period_end,
            status='approved',
        ).annotate(
            month=TruncMonth('expense_date')
        ).values('month').annotate(
            total=Sum('total_amount'),
            vat=Sum('vat_amount'),
        ).order_by('month')
        
        monthly_income = IncomeRecord.objects.filter(
            received_date__gte=period_start,
            received_date__lte=period_end,
        ).annotate(
            month=TruncMonth('received_date')
        ).values('month').annotate(
            total=Sum('total_amount'),
            vat=Sum('vat_amount'),
        ).order_by('month')
        
        # Merge monthly data
        monthly = {}
        for row in monthly_expenses:
            key = row['month'].strftime('%Y-%m')
            monthly[key] = {
                'month': key,
                'expenses': float(row['total']),
                'expenses_vat': float(row['vat']),
                'income': 0,
                'income_vat': 0,
            }
        for row in monthly_income:
            key = row['month'].strftime('%Y-%m')
            if key not in monthly:
                monthly[key] = {'month': key, 'expenses': 0, 'expenses_vat': 0, 'income': 0, 'income_vat': 0}
            monthly[key]['income'] = float(row['total'])
            monthly[key]['income_vat'] = float(row['vat'])
        
        return Response({
            'period_start': period_start.isoformat(),
            'period_end': period_end.isoformat(),
            'year': year,
            'total_income': str(total_income),
            'total_expenses': str(total_expenses),
            'net_profit': str(total_income - total_expenses),
            'total_income_excl_vat': str(income_totals['total_excl_vat'] or Decimal('0.00')),
            'total_expenses_excl_vat': str(expense_totals['total_excl_vat'] or Decimal('0.00')),
            'total_vat_collected': str(total_vat_collected),
            'total_vat_paid': str(total_vat_paid),
            'vat_due': str(total_vat_collected - total_vat_paid),
            'expenses_by_category': [
                {
                    'category': row['category__name'],
                    'code': row['category__code'],
                    'color': row['category__color'],
                    'total': str(row['total'] or Decimal('0.00')),
                }
                for row in by_category
            ],
            'monthly_breakdown': sorted(monthly.values(), key=lambda x: x['month']),
        })
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """
        Export expenses and income to Excel for Aangifte.
        
        Generates a professional Excel file with:
        - Sheet 1: All Expenses
        - Sheet 2: All Income
        - Sheet 3: Summary by Category & Month
        - Sheet 4: BTW Aangifte (Box 1a-5b format)
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            return Response(
                {'error': 'openpyxl is required for Excel export. Install with: pip install openpyxl'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        year = int(request.query_params.get('year', date.today().year))
        quarter = request.query_params.get('quarter')
        
        # Build date range
        if quarter:
            q = int(quarter)
            period_start = date(year, (q - 1) * 3 + 1, 1)
            if q == 4:
                period_end = date(year, 12, 31)
            else:
                period_end = date(year, q * 3 + 1, 1) - timedelta(days=1)
            period_label = f"Q{q} {year}"
        else:
            period_start = date(year, 1, 1)
            period_end = date(year, 12, 31)
            period_label = str(year)
        
        wb = openpyxl.Workbook()
        
        # Styles
        header_font = Font(bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        title_font = Font(bold=True, size=14, color='1E3A5F')
        currency_fmt = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )
        
        # ── Sheet 1: Expenses ──
        ws1 = wb.active
        ws1.title = "Uitgaven (Expenses)"
        ws1['A1'] = f"Uitgaven — {period_label}"
        ws1['A1'].font = title_font
        
        headers = ['Datum', 'Leverancier', 'Omschrijving', 'Categorie', 'Bedrag excl. BTW', 'BTW %', 'BTW Bedrag', 'Totaal incl. BTW', 'Betaalwijze', 'Referentie']
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        expenses = Expense.objects.filter(
            expense_date__gte=period_start,
            expense_date__lte=period_end,
            status='approved',
        ).select_related('category').order_by('expense_date')
        
        for row_idx, exp in enumerate(expenses, 4):
            ws1.cell(row=row_idx, column=1, value=exp.expense_date.strftime('%d-%m-%Y'))
            ws1.cell(row=row_idx, column=2, value=exp.vendor_name)
            ws1.cell(row=row_idx, column=3, value=exp.description)
            ws1.cell(row=row_idx, column=4, value=exp.category.name)
            ws1.cell(row=row_idx, column=5, value=float(exp.amount_excl_vat)).number_format = currency_fmt
            ws1.cell(row=row_idx, column=6, value=float(exp.vat_rate))
            ws1.cell(row=row_idx, column=7, value=float(exp.vat_amount)).number_format = currency_fmt
            ws1.cell(row=row_idx, column=8, value=float(exp.total_amount)).number_format = currency_fmt
            ws1.cell(row=row_idx, column=9, value=exp.get_payment_method_display())
            ws1.cell(row=row_idx, column=10, value=exp.reference_number)
        
        # Auto-fit columns
        for col in range(1, 11):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # ── Sheet 2: Income ──
        ws2 = wb.create_sheet("Inkomsten (Income)")
        ws2['A1'] = f"Inkomsten — {period_label}"
        ws2['A1'].font = title_font
        
        inc_headers = ['Datum', 'Bron', 'Omschrijving', 'Betaler', 'Bedrag excl. BTW', 'BTW Bedrag', 'Totaal incl. BTW', 'Factuurnr.']
        for col, h in enumerate(inc_headers, 1):
            cell = ws2.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        incomes = IncomeRecord.objects.filter(
            received_date__gte=period_start,
            received_date__lte=period_end,
        ).order_by('received_date')
        
        for row_idx, inc in enumerate(incomes, 4):
            ws2.cell(row=row_idx, column=1, value=inc.received_date.strftime('%d-%m-%Y'))
            ws2.cell(row=row_idx, column=2, value=inc.get_source_display())
            ws2.cell(row=row_idx, column=3, value=inc.description)
            ws2.cell(row=row_idx, column=4, value=inc.payer_name)
            ws2.cell(row=row_idx, column=5, value=float(inc.amount_excl_vat)).number_format = currency_fmt
            ws2.cell(row=row_idx, column=6, value=float(inc.vat_amount)).number_format = currency_fmt
            ws2.cell(row=row_idx, column=7, value=float(inc.total_amount)).number_format = currency_fmt
            ws2.cell(row=row_idx, column=8, value=inc.customer_invoice.invoice_number if inc.customer_invoice else '')
        
        for col in range(1, 9):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
        
        # ── Sheet 3: Summary ──
        ws3 = wb.create_sheet("Overzicht (Summary)")
        ws3['A1'] = f"Financieel Overzicht — {period_label}"
        ws3['A1'].font = title_font
        
        total_exp = expenses.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
        total_inc = incomes.aggregate(t=Sum('total_amount'))['t'] or Decimal('0')
        total_exp_vat = expenses.aggregate(t=Sum('vat_amount'))['t'] or Decimal('0')
        total_inc_vat = incomes.aggregate(t=Sum('vat_amount'))['t'] or Decimal('0')
        
        ws3.cell(row=3, column=1, value="Totaal Inkomsten").font = Font(bold=True)
        ws3.cell(row=3, column=2, value=float(total_inc)).number_format = currency_fmt
        ws3.cell(row=4, column=1, value="Totaal Uitgaven").font = Font(bold=True)
        ws3.cell(row=4, column=2, value=float(total_exp)).number_format = currency_fmt
        ws3.cell(row=5, column=1, value="Netto Winst/Verlies").font = Font(bold=True, color='1E3A5F')
        ws3.cell(row=5, column=2, value=float(total_inc - total_exp)).number_format = currency_fmt
        
        ws3.cell(row=7, column=1, value="BTW Overzicht").font = Font(bold=True, size=12, color='1E3A5F')
        ws3.cell(row=8, column=1, value="BTW Ontvangen (Verkopen)")
        ws3.cell(row=8, column=2, value=float(total_inc_vat)).number_format = currency_fmt
        ws3.cell(row=9, column=1, value="BTW Betaald (Inkopen)")
        ws3.cell(row=9, column=2, value=float(total_exp_vat)).number_format = currency_fmt
        ws3.cell(row=10, column=1, value="BTW Af te dragen").font = Font(bold=True)
        ws3.cell(row=10, column=2, value=float(total_inc_vat - total_exp_vat)).number_format = currency_fmt
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 18
        
        # ── Sheet 4: BTW Aangifte ──
        ws4 = wb.create_sheet("BTW Aangifte")
        ws4['A1'] = f"BTW Aangifte — {period_label}"
        ws4['A1'].font = title_font
        
        # Separate expenses by VAT rate
        exp_21 = expenses.filter(vat_rate=Decimal('21.00'))
        exp_9 = expenses.filter(vat_rate=Decimal('9.00'))
        exp_0 = expenses.filter(vat_rate=Decimal('0.00'))
        
        ws4.cell(row=3, column=1, value="Rubriek").font = header_font
        ws4['A3'].fill = header_fill
        ws4.cell(row=3, column=2, value="Omschrijving").font = header_font
        ws4['B3'].fill = header_fill
        ws4.cell(row=3, column=3, value="Bedrag").font = header_font
        ws4['C3'].fill = header_fill
        ws4.cell(row=3, column=4, value="BTW").font = header_font
        ws4['D3'].fill = header_fill
        
        # Box 1a: Leveringen/diensten belast met 21%
        inc_excl = incomes.aggregate(t=Sum('amount_excl_vat'))['t'] or Decimal('0')
        ws4.cell(row=4, column=1, value="1a")
        ws4.cell(row=4, column=2, value="Leveringen/diensten belast met 21%")
        ws4.cell(row=4, column=3, value=float(inc_excl)).number_format = currency_fmt
        ws4.cell(row=4, column=4, value=float(total_inc_vat)).number_format = currency_fmt
        
        # Box 5a: Verschuldigde omzetbelasting
        ws4.cell(row=6, column=1, value="5a")
        ws4.cell(row=6, column=2, value="Verschuldigde omzetbelasting")
        ws4.cell(row=6, column=4, value=float(total_inc_vat)).number_format = currency_fmt
        
        # Box 5b: Voorbelasting
        ws4.cell(row=7, column=1, value="5b")
        ws4.cell(row=7, column=2, value="Voorbelasting (BTW op inkopen)")
        ws4.cell(row=7, column=4, value=float(total_exp_vat)).number_format = currency_fmt
        
        # Box 5g: Af te dragen
        ws4.cell(row=9, column=1, value="5g").font = Font(bold=True, size=12)
        ws4.cell(row=9, column=2, value="Af te dragen / Terug te vragen").font = Font(bold=True, size=12)
        ws4.cell(row=9, column=4, value=float(total_inc_vat - total_exp_vat)).number_format = currency_fmt
        ws4['D9'].font = Font(bold=True, size=12, color='1E3A5F')
        
        ws4.column_dimensions['A'].width = 10
        ws4.column_dimensions['B'].width = 40
        ws4.column_dimensions['C'].width = 18
        ws4.column_dimensions['D'].width = 18
        
        # Write to response
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Aangifte_{period_label.replace(' ', '_')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    
    @action(detail=False, methods=['post'])
    def ocr_extract(self, request):
        """
        Extract data from an uploaded receipt using OCR.
        
        Accepts a file upload and returns extracted fields:
        vendor_name, amount, vat_rate, date, description.
        
        For now, uses a simple heuristic approach.
        Can be enhanced with Google Cloud Vision or Tesseract later.
        """
        receipt = request.FILES.get('receipt')
        if not receipt:
            return Response(
                {'error': 'No receipt file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # For now, return a placeholder response
        # TODO: Integrate with Google Cloud Vision API or Tesseract
        return Response({
            'status': 'success',
            'extracted': {
                'vendor_name': '',
                'amount_excl_vat': '',
                'vat_rate': '21.00',
                'expense_date': date.today().isoformat(),
                'description': receipt.name.rsplit('.', 1)[0] if receipt.name else '',
                'reference_number': '',
            },
            'confidence': 0,
            'message': 'OCR extraction placeholder. Upload saved — please fill in details manually.',
        })


# =============================================================================
# INCOME RECORD VIEWSET
# =============================================================================

class IncomeRecordViewSet(viewsets.ModelViewSet):
    """CRUD for income records."""
    
    queryset = IncomeRecord.objects.order_by('-received_date')
    permission_classes = [IsAdmin]
    
    def get_serializer_class(self):
        if self.action == 'list':
            return IncomeRecordListSerializer
        if self.action in ['create', 'update', 'partial_update']:
            return IncomeRecordCreateSerializer
        return IncomeRecordDetailSerializer
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        year = self.request.query_params.get('year')
        if year:
            qs = qs.filter(received_date__year=int(year))
        
        source = self.request.query_params.get('source')
        if source:
            qs = qs.filter(source=source)
        
        search = self.request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(description__icontains=search) |
                Q(payer_name__icontains=search)
            )
        
        return qs
